def StabDia(NatFreq, FRF, FRF_est, Freq, Order, sensor, test_series, iters, recon = 'yes', algo ='', ):
    import matplotlib.pyplot as plt 
    import numpy as np
    import openpyxl
    # Create figure and subplot
    fig, host = plt.subplots(figsize=(15,10))   
    plt.grid()  
    par1 = host.twinx()
    host.set_xlabel("Frequency[Hz]")
    host.set_ylabel("FRF")
    par1.set_ylabel("Model Order")
    par1.set_yticks(Order)
    color1 = plt.cm.viridis(0)
    if algo =='P':
        for i in range(np.shape(FRF)[2]):
            FRFs = [abs(frf) for frf in FRF[0,:,i]]
            host.semilogy(Freq, FRFs)
    else:
            FRF = [abs(frf) for frf in FRF]
            host.semilogy(Freq, FRF, color=color1, label="FRF")
        
    if recon == 'yes':
        for i, j in zip(FRF_est, Order):
            k = [abs(frf) for frf in i]
            host.semilogy(Freq, k, label =j)
    elif recon == 'no':
        pass
    else:
        print('recon = yes/no')
    host.legend(loc="upper left")
    for w_n, N in zip(NatFreq,Order):
        N= [N for i in range(len(w_n))]
        #par1.plot(w_n, N,"+", markersize=5)
        if len(w_n) > 0:
            plt.vlines(w_n, ymin=0, ymax=1, ls='--', lw=2, label='wn')        
    plt.title(sensor)
    
    def onclick(event):
    # Get the x-coordinate of the click
        x = event.xdata
        col = 28
        if x is not None:
            # Round the x-coordinate to the nearest natural frequency
            nearest_freq = min(w_n, key=lambda f: abs(f-x))
            print('Selected frequency:', nearest_freq)
    
            # Write the selected frequency to an Excel file
            wb = openpyxl.load_workbook('selected_frequency.xlsx')
            sheet_name = test_series+str(iters[0])
            sheet_names = wb.sheetnames
            if sheet_name in sheet_names:
                ws = wb[sheet_name]
                #ws = wb.active
                # Find the first empty row
                row = 1
                while ws.cell(row=row, column=col).value is not None:
                    row += 1
            else:
                ws = wb.create_sheet(sheet_name)
                row = 1
                while ws.cell(row=row, column=col).value is not None:
                    row += 1
            # Write the frequency to the next empty row
            ws.cell(row=row, column=col-1).value = sensor
            ws.cell(row=row, column=col).value = nearest_freq
            wb.save('selected_frequency.xlsx')
    # Connect the onclick function to the plot
    cid = fig.canvas.mpl_connect('button_press_event', onclick)
    plt.show()
    return fig